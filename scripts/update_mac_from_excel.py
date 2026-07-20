"""
Aggiorna docs/data/mac_anagrafica.json dalla sheet "Anagrafica Articoli"
del file Excel modello MACAO (05_Modello_Standard_Cost_MACAO.xlsx).

Uso:
  python scripts/update_mac_from_excel.py <percorso_excel>

Esempio:
  python scripts/update_mac_from_excel.py "C:/Users/.../05_Modello_Standard_Cost_MACAO.xlsx"

Colonne Anagrafica Articoli (A:V):
  A  = MACAO no          → id, nHK
  B  = BV no             → code
  C  = ifbitem           → ifbItem
  D  = description       → description
  E  = sectiondescription→ category
  F  = producttype       → producttype, temperature
  G  = vendorname        → vendorName
  H  = standardcost(HKD) → standardCostHkd
  I  = hoff              → isHoff
  L  = sales UOM BV      → hkUom
  M  = sales UOM MACAO   → uom
  N  = pcsgrossweight MACAO → pcsgrossweight  (usato dal modello Excel)
  R  = quantityxpackaging MACAO → qtyPerBox   (usato dal modello Excel)
  V  = vendorname2       → vendorName2
"""
import sys, json, re
import openpyxl
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
OUT_PATH   = SCRIPT_DIR.parent / "docs" / "data" / "mac_anagrafica.json"

UOM_MAP  = {"pcs": "PCS", "pz": "PCS", "piece": "PCS", "box": "BOX", "ctn": "BOX",
            "cartone": "BOX", "kg": "KG", "kgs": "KG", "gr": "GR", "gram": "GR"}
TEMP_MAP = {"dry": "DRY", "fresh": "FRESH", "frozen": "FROZEN",
            "ambient": "DRY", "refrigerated": "FRESH"}

def norm_uom(v):
    s = str(v or "").strip().lower()
    return UOM_MAP.get(s, str(v or "").upper() if v else "KG")

def norm_temp(v):
    s = str(v or "").strip().lower()
    return TEMP_MAP.get(s, str(v or "DRY").upper())

def looks_like_code(v):
    """Filtra righe non-prodotto (note filtri, header duplicati, ecc.)"""
    if not v:
        return False
    s = str(v).strip()
    if len(s) > 40:
        return False
    if "filtri" in s.lower() or "applicati" in s.lower():
        return False
    return bool(re.search(r"[A-Za-z0-9]", s))


def main():
    if len(sys.argv) < 2:
        print("Uso: python scripts/update_mac_from_excel.py <percorso_excel>")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f"ERRORE: file non trovato: {excel_path}")
        sys.exit(1)

    print(f"Leggo Anagrafica da: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    if "Anagrafica Articoli" not in wb.sheetnames:
        print(f"ERRORE: sheet 'Anagrafica Articoli' non trovata. Sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["Anagrafica Articoli"]
    products = []
    skipped  = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        mac_no = row[0]
        if not looks_like_code(mac_no):
            skipped += 1
            continue

        mac_no  = str(mac_no).strip()
        bv_no   = str(row[1]).strip() if row[1] else mac_no
        ifb     = str(row[2]).strip() if row[2] and row[2] != 0 else None
        desc    = str(row[3] or "").strip()
        sect    = str(row[4] or "").strip()
        ptype   = norm_temp(row[5])
        vendor  = str(row[6] or "BRIGHT VIEW TRADING HK LTD").strip()
        sc_hkd  = float(row[7] or 0)
        hoff    = bool(row[8])

        # UOM (indici 0-based: L=col 11, M=col 12)
        uom_bv  = norm_uom(row[11]) if len(row) > 11 else "KG"
        uom_mac = norm_uom(row[12]) if len(row) > 12 else uom_bv

        # pcsgrossweight MACAO (col N = indice 13)
        pcsgross = float(row[13] or 0) if len(row) > 13 else 0.0

        # quantityxpackaging MACAO (col R = indice 17)
        qty_mac  = float(row[17] or 1) if len(row) > 17 else 1.0

        # vendorname2 (col V = indice 21)
        vendor2 = str(row[21]).strip() if len(row) > 21 and row[21] else None

        products.append({
            "id":               mac_no,
            "nHK":              mac_no,
            "code":             bv_no,
            **({"ifbItem": ifb} if ifb else {}),
            "description":      desc,
            "category":         sect,
            "producttype":      ptype,
            "temperature":      ptype,
            "vendorName":       vendor,
            **({"vendorName2": vendor2} if vendor2 else {}),
            "uom":              uom_mac,
            "hkUom":            uom_bv,
            "pcsgrossweight":   pcsgross,
            "qtyPerBox":        qty_mac,
            "standardCostHkd":  round(sc_hkd, 5),
            "isHoff":           hoff,
            "active":           True,
        })

    print(f"  {len(products)} prodotti letti, {skipped} righe saltate")
    if products:
        hoff_n = sum(1 for p in products if p["isHoff"])
        sc_n   = sum(1 for p in products if p["standardCostHkd"] > 0)
        print(f"  isHoff=True: {hoff_n}, standardCostHkd > 0: {sc_n}")
        print(f"  Esempio: {json.dumps(products[0], ensure_ascii=False)}")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Scritto {len(products)} prodotti in {OUT_PATH}")


if __name__ == "__main__":
    main()
